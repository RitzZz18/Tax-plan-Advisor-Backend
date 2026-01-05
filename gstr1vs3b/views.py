import uuid
import requests
from datetime import datetime, timedelta

from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
import calendar
from datetime import datetime, date

from .models import GSTSession, ReconciliationReport
from gst_auth.utils import get_valid_session


# ---------------------------------------------------------
# 🔧 Utility: Safe API request wrapper
# ---------------------------------------------------------
def safe_api_call(method, url, **kwargs):
    """Unified request handler for cleaner code."""
    try:
        kwargs["timeout"] = 20
        res = requests.request(method, url, **kwargs)
        try:
            data = res.json()
        except:
            data = {}

        return res.status_code, data

    except requests.Timeout:
        return 504, {"error": "timeout"}

    except requests.RequestException:
        return 503, {"error": "connection_failed"}

    except Exception:
        return 500, {"error": "internal_error"}


# ---------------------------------------------------------
# 🔹 1. GENERATE OTP
# ---------------------------------------------------------
@api_view(['POST'])
@permission_classes([AllowAny])
def generate_otp(request):
    username = request.data.get("username")
    gstin = request.data.get("gstin")

    # Input Validation
    if not username or not username.strip():
        return Response({"error": "Username required"}, status=400)

    if not gstin or len(gstin) != 15:
        return Response({"error": "GSTIN must be 15 characters"}, status=400)

    # Step 1 → Authenticate
    status_code, auth_data = safe_api_call(
        "POST",
        "https://api.sandbox.co.in/authenticate",
        headers={
            "x-api-key": settings.SANDBOX_API_KEY,
            "x-api-secret": settings.SANDBOX_API_SECRET
        }
    )

    if status_code != 200:
        return Response({"error": "Failed to authenticate"}, status=500)

    access_token = auth_data.get("data", {}).get("access_token")
    if not access_token:
        return Response({"error": "Invalid token from GST API"}, status=500)

    # Step 2 → Send OTP
    status_code, otp_data = safe_api_call(
        "POST",
        "https://api.sandbox.co.in/gst/compliance/tax-payer/otp",
        json={"username": username, "gstin": gstin},
        headers={
            "x-source": "primary",
            "x-api-version": "1.0.0",
            "Authorization": access_token,
            "x-api-key": settings.SANDBOX_API_KEY,
            "Content-Type": "application/json"
        }
    )

    data = otp_data.get("data", {})

    if data.get("status_cd") == "0":
        return Response({
            "error": data.get("message", "OTP failed"),
            "error_code": data.get("error", {}).get("error_cd", "")
        }, status=400)

    # Create DB session
    gst_session = GSTSession.objects.create(
        session_id=uuid.uuid4(),
        username=username,
        gstin=gstin,
        access_token=access_token
    )

    # Auto-clean old sessions
    GSTSession.objects.filter(
        created_at__lt=datetime.now() - timedelta(hours=24)
    ).delete()

    return Response({
        "message": "OTP sent successfully",
        "session_id": str(gst_session.session_id)
    })


# ---------------------------------------------------------
# 🔹 2. VERIFY OTP
# ---------------------------------------------------------
@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    otp = request.data.get("otp")
    session_id = request.data.get("session_id")

    if not otp or not otp.strip():
        return Response({"error": "OTP required"}, status=400)

    if not session_id:
        return Response({"error": "Session ID required"}, status=400)

    try:
        gst_session = GSTSession.objects.get(session_id=session_id)
    except GSTSession.DoesNotExist:
        return Response({"error": "Session expired"}, status=400)

    # Verify OTP
    status_code, verify_data = safe_api_call(
        "POST",
        "https://api.sandbox.co.in/gst/compliance/tax-payer/otp/verify",
        json={"username": gst_session.username, "gstin": gst_session.gstin},
        params={"otp": otp},
        headers={
            "x-source": "primary",
            "x-api-version": "1.0.0",
            "Authorization": gst_session.access_token,
            "x-api-key": settings.SANDBOX_API_KEY,
            "Content-Type": "application/json"
        }
    )

    data = verify_data.get("data", {})
    taxpayer_token = data.get("access_token")
    if data.get("status_cd") == "0" or not taxpayer_token:
        return Response({"error": "OTP verification failed"}, status=400)

    gst_session.taxpayer_token = taxpayer_token
    gst_session.save(update_fields=["taxpayer_token", "updated_at"])

    return Response({
        "message": "OTP verified successfully",
        "session_id": str(gst_session.session_id)
    })


# =====================================================
# 🔧 HELPER: SAFE VALUE EXTRACTOR
# =====================================================
def get_val(source_dict, key, field):
    """
    Tries to fetch value from:
    1. source_dict[key]['subtotal'][field] (Auto-Pop style)
    2. source_dict[key][field] (Filed style)
    3. Returns 0.0 if not found
    """
    if not source_dict:
        return 0.0
        
    section = source_dict.get(key, {})
    if not section:
        return 0.0

    # Try Auto-Populated Structure (nested in subtotal)
    if "subtotal" in section:
        return float(section["subtotal"].get(field, 0) or 0)
    
    # Try Filed Structure (direct value)
    return float(section.get(field, 0) or 0)


# =====================================================
# 1. FETCH AUTO LIABILITY (The Govt Data)
# =====================================================
def fetch_auto_liability(year, month, headers):
    url = f"https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-3b/{year}/{month:02d}/auto-liability-calc"
    status, data = safe_api_call("GET", url, headers=headers)

    if status != 200:
        return None

    # Navigate to sup_details with extreme caution
    # Path: data -> data -> r3bautopop -> liabitc -> sup_details
    try:
        sup_details = (
            data.get("data", {})
                .get("data", {})
                .get("r3bautopop", {})
                .get("liabitc", {})
                .get("sup_details", {})
        )
    except:
        return None

    if not sup_details:
        return None

    # Use the helper to safely grab values regardless of structure
    return {
        # 3.1.a Standard
        "tx": get_val(sup_details, "osup_3_1a", "txval"),
        "igst": get_val(sup_details, "osup_3_1a", "iamt"),
        "cgst": get_val(sup_details, "osup_3_1a", "camt"),
        "sgst": get_val(sup_details, "osup_3_1a", "samt"),
        
        # 3.1.b Exports
        "exp_tx": get_val(sup_details, "osup_3_1b", "txval"),
        "exp_igst": get_val(sup_details, "osup_3_1b", "iamt"),

        # 3.1.c Nil Rated
        "nil_tx": get_val(sup_details, "osup_3_1c", "txval"),

        # 3.1.d RCM (Liability)
        # "rcm_tx": get_val(sup_details, "isup_3_1d", "txval"),
        # "rcm_igst": get_val(sup_details, "isup_3_1d", "iamt"),
        # "rcm_cgst": get_val(sup_details, "isup_3_1d", "camt"),
        # "rcm_sgst": get_val(sup_details, "isup_3_1d", "samt"),

        # 3.1.e Non-GST
        "nongst_tx": get_val(sup_details, "osup_3_1e", "txval"),
    }

# =====================================================
# 2. FETCH FILED GSTR-3B (The User Data) + ITC Data
# =====================================================
def fetch_filed_3b(year, month, headers):
    url = f"https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-3b/{year}/{month:02d}"
    status, data = safe_api_call("GET", url, headers=headers)

    if status != 200:
        return None

    # Navigate to full 3B data
    # Path: data -> data
    try:
        gstr3b_data = data.get("data", {}).get("data", {})
        sup_details = gstr3b_data.get("sup_details", {})
        itc_elg = gstr3b_data.get("itc_elg", {})
    except:
        return None
        
    if not sup_details:
        return None

    # MAPPING KEYS:
    # Standard -> osup_det
    # Exports  -> osup_zero
    # Nil      -> osup_nil_exmp
    # RCM      -> isup_rev (Note: Key is distinct from Auto)
    # Non-GST  -> osup_nongst

    # Extract ITC data from itc_elg (Table 4A)
    # Table 4A - ITC Available breakdown:
    # - ty = "IMPG" (Imports of Goods) - NOT in 2B
    # - ty = "IMPS" (Imports of Services/RCM) - NOT in 2B
    # - ty = "ISRC" (Inward supplies liable to RCM) - NOT in 2B
    # - ty = "ISD" (Input from ISD) - May be in 2B
    # - ty = "OTH" (All other ITC) - Should match 2B B2B
    
    itc_avl = itc_elg.get("itc_avl", [])
    
    # Total ITC claimed
    itc_total_igst = 0
    itc_total_cgst = 0
    itc_total_sgst = 0
    itc_total_cess = 0
    
    # ITC from RCM/Imports (NOT in 2B) - for adjustment
    itc_rcm_igst = 0
    itc_rcm_cgst = 0
    itc_rcm_sgst = 0
    itc_rcm_cess = 0
    
    for item in itc_avl:
        igst = float(item.get("iamt", 0) or 0)
        cgst = float(item.get("camt", 0) or 0)
        sgst = float(item.get("samt", 0) or 0)
        cess = float(item.get("csamt", 0) or 0)
        
        # Add to total
        itc_total_igst += igst
        itc_total_cgst += cgst
        itc_total_sgst += sgst
        itc_total_cess += cess
        
        # Check type - RCM/Imports are NOT in 2B
        item_type = item.get("ty", "")
        if item_type in ["IMPG", "IMPS", "ISRC"]:
            itc_rcm_igst += igst
            itc_rcm_cgst += cgst
            itc_rcm_sgst += sgst
            itc_rcm_cess += cess

    return {
        # Standard (Section 3.1.a)
        "tx": get_val(sup_details, "osup_det", "txval"),
        "igst": get_val(sup_details, "osup_det", "iamt"),
        "cgst": get_val(sup_details, "osup_det", "camt"),
        "sgst": get_val(sup_details, "osup_det", "samt"),
        
        # Exports (Section 3.1.b)
        "exp_tx": get_val(sup_details, "osup_zero", "txval"),
        "exp_igst": get_val(sup_details, "osup_zero", "iamt"),

        # Nil (Section 3.1.c)
        "nil_tx": get_val(sup_details, "osup_nil_exmp", "txval"),

        # Non-GST (Section 3.1.e)
        "nongst_tx": get_val(sup_details, "osup_nongst", "txval"),
        
        # Total ITC Claimed in 3B (Table 4A)
        "itc_igst": itc_total_igst,
        "itc_cgst": itc_total_cgst,
        "itc_sgst": itc_total_sgst,
        "itc_cess": itc_total_cess,
        
        # RCM/Imports ITC (NOT in 2B - for adjustment)
        "itc_rcm_igst": itc_rcm_igst,
        "itc_rcm_cgst": itc_rcm_cgst,
        "itc_rcm_sgst": itc_rcm_sgst,
        "itc_rcm_cess": itc_rcm_cess,
    }


# =====================================================
# 3. FETCH GSTR-2B DATA (Purchase ITC from suppliers)
# =====================================================
def fetch_2b_data(year, month, headers):
    """
    Fetch GSTR-2B ITC data from Sandbox API.
    GSTR-2B contains eligible ITC from supplier invoices.
    
    Actual API Response Structure:
    data.data.data.itcsumm.itcavl.nonrevsup = {
        "sgst": 9442.62,
        "cgst": 9442.62,
        "igst": 5885.22,
        "cess": 0.0,
        "b2b": { ... detailed values ... }
    }
    """
    url = f"https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-2b/{year}/{month:02d}"
    status, data = safe_api_call("GET", url, headers=headers)

    if status != 200:
        return None

    # Navigate to ITC summary
    # Path: data -> data -> data -> itcsumm -> itcavl
    try:
        # API returns: data.data.data.itcsumm (3 levels of 'data')
        inner_data = data.get("data", {}).get("data", {}).get("data", {})
        itcsumm = inner_data.get("itcsumm", {})
        itcavl = itcsumm.get("itcavl", {})
        
        # Get ITC from non-reverse charge supplies (main B2B ITC)
        nonrevsup = itcavl.get("nonrevsup", {})
        
        # Also get other sources if available
        othersup = itcavl.get("othersup", {})
        
    except Exception as e:
        print(f"Error parsing 2B data: {e}")
        return None

    # Helper to extract ITC values using correct keys: igst, cgst, sgst, cess
    def get_itc_vals(section):
        if not section:
            return 0, 0, 0, 0
        return (
            float(section.get("igst", 0) or 0),
            float(section.get("cgst", 0) or 0),
            float(section.get("sgst", 0) or 0),
            float(section.get("cess", 0) or 0)
        )
    
    # Non-reverse charge supplies (main B2B ITC) - use top-level values
    nr_igst, nr_cgst, nr_sgst, nr_cess = get_itc_vals(nonrevsup)
    
    # Other supplies (includes credit notes etc)
    oth_igst, oth_cgst, oth_sgst, oth_cess = get_itc_vals(othersup)
    
    # Total ITC available as per 2B
    total_igst = nr_igst + oth_igst
    total_cgst = nr_cgst + oth_cgst
    total_sgst = nr_sgst + oth_sgst
    total_cess = nr_cess + oth_cess

    return {
        "itc_igst": total_igst,
        "itc_cgst": total_cgst,
        "itc_sgst": total_sgst,
        "itc_cess": total_cess,
        # Also include individual components for detailed view
        "b2b_igst": nr_igst,
        "b2b_cgst": nr_cgst,
        "b2b_sgst": nr_sgst,
        "oth_igst": oth_igst,
        "oth_cgst": oth_cgst,
        "oth_sgst": oth_sgst,
    }

# =====================================================
# 4. RECONCILE MONTH (The Logic) - Now includes 2B data
# =====================================================
def reconcile_month(year, month, token):
    headers = {
        "Authorization": token,
        "x-api-key": settings.SANDBOX_API_KEY,
        "x-api-version": "1.0.0"
    }

    # Fetch all data sources
    auto = fetch_auto_liability(year, month, headers)  # GSTR-1 auto-populated
    filed = fetch_filed_3b(year, month, headers)       # Filed GSTR-3B (includes ITC)
    gstr2b = fetch_2b_data(year, month, headers)       # GSTR-2B (purchase ITC)

    if not auto or not filed:
        return None

    # Sales reconciliation (GSTR-1 vs GSTR-3B)
    sales_diff = (
        abs(auto["tx"] - filed["tx"]) +
        abs(auto["igst"] - filed["igst"]) +
        abs(auto["cgst"] - filed["cgst"]) +
        abs(auto["sgst"] - filed["sgst"]) +
        abs(auto["exp_tx"] - filed["exp_tx"]) +
        abs(auto["nongst_tx"] - filed["nongst_tx"])
    )

    # Purchase ITC reconciliation (GSTR-2B vs GSTR-3B ITC claimed)
    # Note: 3B ITC includes RCM/Imports which are NOT in 2B
    # So we compare: 2B vs (3B - RCM) for "adjusted" match
    if gstr2b:
        # Gross difference (will mismatch due to RCM)
        itc_diff_gross = (
            abs(gstr2b["itc_igst"] - filed["itc_igst"]) +
            abs(gstr2b["itc_cgst"] - filed["itc_cgst"]) +
            abs(gstr2b["itc_sgst"] - filed["itc_sgst"]) +
            abs(gstr2b["itc_cess"] - filed["itc_cess"])
        )
        
        # Adjusted difference: 2B vs (3B - RCM)
        # This should match if all B2B ITC is properly claimed
        g3_adj_igst = filed["itc_igst"] - filed.get("itc_rcm_igst", 0)
        g3_adj_cgst = filed["itc_cgst"] - filed.get("itc_rcm_cgst", 0)
        g3_adj_sgst = filed["itc_sgst"] - filed.get("itc_rcm_sgst", 0)
        g3_adj_cess = filed["itc_cess"] - filed.get("itc_rcm_cess", 0)
        
        itc_diff_adj = (
            abs(gstr2b["itc_igst"] - g3_adj_igst) +
            abs(gstr2b["itc_cgst"] - g3_adj_cgst) +
            abs(gstr2b["itc_sgst"] - g3_adj_sgst) +
            abs(gstr2b["itc_cess"] - g3_adj_cess)
        )
        
        # Smart status logic:
        # 1. If 3B_adj > 2B for any head → RISK (claiming more than available)
        # 2. Else if 3B_adj ≈ 2B and RCM exists → RECONCILED (RCM excluded)
        # 3. Else → PARTIAL CLAIMED (2B > 3B_adj, not claiming full eligible)
        
        has_rcm = (filed.get("itc_rcm_igst", 0) + filed.get("itc_rcm_cgst", 0) + 
                   filed.get("itc_rcm_sgst", 0) + filed.get("itc_rcm_cess", 0)) > 0
        
        # Check if any adjusted 3B exceeds 2B (RISK scenario)
        excess_igst = g3_adj_igst - gstr2b["itc_igst"]
        excess_cgst = g3_adj_cgst - gstr2b["itc_cgst"]
        excess_sgst = g3_adj_sgst - gstr2b["itc_sgst"]
        excess_cess = g3_adj_cess - gstr2b["itc_cess"]
        
        has_excess = (excess_igst > 5 or excess_cgst > 5 or excess_sgst > 5 or excess_cess > 5)
        
        if has_excess:
            itc_status = "RISK"  # Claiming more than available in 2B
        elif itc_diff_adj < 5.0:
            itc_status = "RECONCILED" if has_rcm else "MATCH"  # Perfect match
        else:
            itc_status = "PARTIAL"  # 2B > 3B_adj, not claiming full eligible
    else:
        itc_status = "NO 2B DATA"
        g3_adj_igst = filed["itc_igst"]
        g3_adj_cgst = filed["itc_cgst"]
        g3_adj_sgst = filed["itc_sgst"]
        g3_adj_cess = filed["itc_cess"]

    return {
        "year": year, "month": month,
        
        # ----- SALES: GSTR-1 vs GSTR-3B -----
        "auto_tx": auto["tx"], "g3_tx": filed["tx"],
        "auto_igst": auto["igst"], "g3_igst": filed["igst"],
        "auto_cgst": auto["cgst"], "g3_cgst": filed["cgst"],
        "auto_sgst": auto["sgst"], "g3_sgst": filed["sgst"],
        
        "auto_exp_tx": auto["exp_tx"], "g3_exp_tx": filed["exp_tx"],
        "auto_exp_igst": auto["exp_igst"], "g3_exp_igst": filed["exp_igst"],
        
        "auto_nil_tx": auto["nil_tx"], "g3_nil_tx": filed["nil_tx"],
        "auto_nongst_tx": auto["nongst_tx"], "g3_nongst_tx": filed["nongst_tx"],
        
        "sales_status": "MATCH" if sales_diff < 5.0 else "MISMATCH",
        
        # ----- PURCHASES: GSTR-2B vs GSTR-3B ITC -----
        # ITC as per GSTR-2B (from suppliers - what's available)
        "g2b_itc_igst": gstr2b["itc_igst"] if gstr2b else 0,
        "g2b_itc_cgst": gstr2b["itc_cgst"] if gstr2b else 0,
        "g2b_itc_sgst": gstr2b["itc_sgst"] if gstr2b else 0,
        "g2b_itc_cess": gstr2b["itc_cess"] if gstr2b else 0,
        
        # Total ITC claimed in GSTR-3B (includes RCM)
        "g3_itc_igst": filed["itc_igst"],
        "g3_itc_cgst": filed["itc_cgst"],
        "g3_itc_sgst": filed["itc_sgst"],
        "g3_itc_cess": filed["itc_cess"],
        
        # RCM/Imports ITC in 3B (NOT in 2B - for adjustment)
        "g3_rcm_igst": filed.get("itc_rcm_igst", 0),
        "g3_rcm_cgst": filed.get("itc_rcm_cgst", 0),
        "g3_rcm_sgst": filed.get("itc_rcm_sgst", 0),
        "g3_rcm_cess": filed.get("itc_rcm_cess", 0),
        
        # 3B ITC after removing RCM (should match 2B)
        "g3_adj_igst": g3_adj_igst,
        "g3_adj_cgst": g3_adj_cgst,
        "g3_adj_sgst": g3_adj_sgst,
        "g3_adj_cess": g3_adj_cess,
        
        "itc_status": itc_status,
        
        # Overall status
        "status": "MATCH" if sales_diff < 5.0 else "MISMATCH"
    }

# =====================================================
# 5. FULL YEAR RECONCILE ENDPOINT (Orchestrator)
# =====================================================

# =====================================================
# RECONCILE ENDPOINT (With Period Selection + 10th of Month Cutoff)
# =====================================================
@api_view(["POST"])
@permission_classes([AllowAny])
def reconcile(request):
    try:
        fy_year = int(request.data.get("fy_year"))
        session_id = request.data.get("session_id")
        
        # New: Period selection parameters
        period_type = request.data.get("period_type", "fy")  # "fy", "quarter", "month"
        period_value = request.data.get("period_value")  # 1-4 for quarter, 1-12 for month

        if not session_id:
            return Response({"error": "Session ID required"}, status=400)

        # Use shared session utility from gst_auth
        session, error = get_valid_session(session_id)
        if error:
            return Response({"error": error}, status=401)

        # ---------------------------------------------------
        # 🗓️ DATE CUTOFF LOGIC
        # ---------------------------------------------------
        today = date.today()
        
        # If today is 1st-10th: Cutoff is 2 months ago (e.g., On Oct 5, show up to Aug)
        # If today is 11th+: Cutoff is 1 month ago (e.g., On Oct 15, show up to Sep)
        if today.day <= 10:
            cutoff_date = (today.replace(day=1) - timedelta(days=45))
        else:
            cutoff_date = (today.replace(day=1) - timedelta(days=15))
        
        cutoff_y, cutoff_m = cutoff_date.year, cutoff_date.month
        # ---------------------------------------------------

        # Generate months list based on period type
        all_fy_months = [(fy_year, m) for m in range(4, 13)] + [(fy_year + 1, m) for m in range(1, 4)]
        
        if period_type == "month" and period_value:
            # Single month: period_value is 1-12 (April=4, May=5, ..., March=3)
            month_num = int(period_value)
            if month_num >= 4:
                months = [(fy_year, month_num)]
            else:
                months = [(fy_year + 1, month_num)]
                
        elif period_type == "quarter" and period_value:
            # Quarter: Q1 = Apr-Jun, Q2 = Jul-Sep, Q3 = Oct-Dec, Q4 = Jan-Mar
            q = int(period_value)
            quarter_months = {
                1: [(fy_year, 4), (fy_year, 5), (fy_year, 6)],
                2: [(fy_year, 7), (fy_year, 8), (fy_year, 9)],
                3: [(fy_year, 10), (fy_year, 11), (fy_year, 12)],
                4: [(fy_year + 1, 1), (fy_year + 1, 2), (fy_year + 1, 3)],
            }
            months = quarter_months.get(q, all_fy_months)
        else:
            # Default: Full FY
            months = all_fy_months

        results = []

        for y, m in months:
            # Skip future months
            if y > cutoff_y or (y == cutoff_y and m > cutoff_m):
                continue 

            res = reconcile_month(y, m, session.taxpayer_token)
            if res:
                results.append(res)
            else:
                # Add placeholder if no data found - includes all fields
                results.append({
                    "year": y, "month": m, 
                    "status": "NO DATA", "sales_status": "NO DATA", "itc_status": "NO DATA",
                    # Sales fields
                    "auto_tx": 0, "g3_tx": 0, "auto_igst": 0, "g3_igst": 0,
                    "auto_cgst": 0, "g3_cgst": 0, "auto_sgst": 0, "g3_sgst": 0,
                    "auto_exp_tx": 0, "g3_exp_tx": 0, "auto_exp_igst": 0, "g3_exp_igst": 0,
                    "auto_nil_tx": 0, "g3_nil_tx": 0, "auto_nongst_tx": 0, "g3_nongst_tx": 0,
                    # 2B ITC fields
                    "g2b_itc_igst": 0, "g2b_itc_cgst": 0, "g2b_itc_sgst": 0, "g2b_itc_cess": 0,
                    "g3_itc_igst": 0, "g3_itc_cgst": 0, "g3_itc_sgst": 0, "g3_itc_cess": 0,
                    # RCM ITC fields
                    "g3_rcm_igst": 0, "g3_rcm_cgst": 0, "g3_rcm_sgst": 0, "g3_rcm_cess": 0,
                    # Adjusted 3B ITC (3B - RCM)
                    "g3_adj_igst": 0, "g3_adj_cgst": 0, "g3_adj_sgst": 0, "g3_adj_cess": 0,
                })

        # Safe Delete & Update
        ReconciliationReport.objects.filter(
            username=session.username, 
            gstin=session.gstin, 
            fy_year=fy_year
        ).delete()

        ReconciliationReport.objects.create(
            username=session.username,
            gstin=session.gstin,
            fy_year=fy_year,
            report_data=results
        )

        return Response({
            "message": "Reconciliation complete",
            "results": results
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# =====================================================
# EXCEL DOWNLOAD (With Sales + Purchases Sheets)
# =====================================================
@api_view(['POST'])
@permission_classes([AllowAny])
def download_excel(request):
    results = request.data.get('results', [])
    username = request.data.get('username', '')
    gstin = request.data.get('gstin', '')
    fy_year = request.data.get('fy_year', '')
    
    wb = Workbook()
    
    # --- Styles ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    month_font = Font(bold=True, color="FFFFFF", size=11)
    diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    itc_header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def create_reco_sheet(ws, title, particulars, subtitle):
        """Helper to create a reconciliation sheet"""
        total_cols = max(len(results) * 4 + 1, 5)
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f"Username: {username} | GSTIN: {gstin} | FY: {fy_year}-{int(fy_year) + 1}"
        ws['A2'].font = Font(bold=True, size=11)
        
        ws['A4'] = "Particular"
        ws['A4'].fill = header_fill
        ws['A4'].font = header_font
        ws['A4'].border = border
        
        # Month Headers
        col = 2
        for data in results:
            month_name = calendar.month_abbr[data['month']] + " " + str(data['year'])
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
            cell = ws.cell(4, col, month_name)
            cell.fill = month_fill
            cell.font = month_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            ws.cell(5, col, subtitle[0]).font = Font(bold=True, size=9)
            ws.cell(5, col+1, subtitle[1]).font = Font(bold=True, size=9)
            ws.cell(5, col+2, "Diff").font = Font(bold=True, size=9)
            col += 4
        
        # Data Rows
        row = 6
        for particular, key_auto, key_filed in particulars:
            ws.cell(row, 1, particular).border = border
            col = 2
            for data in results:
                auto_val = float(data.get(key_auto, 0) or 0)
                filed_val = float(data.get(key_filed, 0) or 0)
                diff = auto_val - filed_val
                
                c1 = ws.cell(row, col, round(auto_val, 2))
                c1.number_format = '#,##0.00'
                c1.border = border
                
                c2 = ws.cell(row, col+1, round(filed_val, 2))
                c2.number_format = '#,##0.00'
                c2.border = border
                
                c3 = ws.cell(row, col+2, round(diff, 2))
                c3.number_format = '#,##0.00'
                c3.border = border
                
                if abs(diff) > 1:
                    c3.fill = diff_fill
                    c3.font = Font(bold=True, color="9C0006")
                else:
                    c3.fill = match_fill
                    c3.font = Font(color="006100")
                
                col += 4
            row += 1
        
        ws.column_dimensions['A'].width = 30
        for i in range(2, col):
            ws.column_dimensions[get_column_letter(i)].width = 18
        ws.freeze_panes = 'B6'
    
    # ========== SHEET 1: Sales (GSTR-1 vs GSTR-3B) ==========
    ws_sales = wb.active
    ws_sales.title = "Sales (R1 vs 3B)"
    
    sales_particulars = [
        ('3.1.a Taxable Value', 'tx1', 'tx3'),
        ('3.1.a IGST', 'ig1', 'ig3'),
        ('3.1.a CGST', 'cg1', 'cg3'),
        ('3.1.a SGST', 'sg1', 'sg3'),
        ('3.1.b Export Taxable', 'exp_tx1', 'exp_tx3'),
        ('3.1.b Export IGST', 'exp_ig1', 'exp_ig3'),
        ('3.1.c Nil/Exempt', 'nil_tx1', 'nil_tx3'),
        ('3.1.e Non-GST', 'ng1', 'ng3'),
    ]
    create_reco_sheet(ws_sales, "GSTR-1 vs GSTR-3B Reconciliation (Sales)", 
                      sales_particulars, ("GSTR-1", "GSTR-3B"))
    
    # ========== SHEET 2: Purchases (GSTR-2B vs GSTR-3B ITC) ==========
    ws_purchases = wb.create_sheet("Purchases (2B vs 3B)")
    
    # Same format as Sales - rows for each tax type, compare 2B vs 3B Adjusted
    # NOTE: Using FRONTEND keys since data comes from frontend mapping
    itc_particulars = [
        ('ITC - IGST', 'itc_2b_igst', 'itc_adj_igst'),
        ('ITC - CGST', 'itc_2b_cgst', 'itc_adj_cgst'),
        ('ITC - SGST', 'itc_2b_sgst', 'itc_adj_sgst'),
        ('ITC - CESS', 'itc_2b_cess', 'itc_adj_cess'),
    ]
    create_reco_sheet(ws_purchases, "GSTR-2B vs GSTR-3B ITC Reconciliation (RCM Adjusted)", 
                      itc_particulars, ("GSTR-2B", "GSTR-3B (Adj)"))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="GSTR_Reconciliation_{gstin}_{fy_year}.xlsx"'
    )
    return response


# =====================================================
# GSTR-3B DETAILS EXCEL DOWNLOAD
# =====================================================
@api_view(['POST'])
@permission_classes([AllowAny])
def download_3b_excel(request):
    """
    Fetch GSTR-3B data for selected month and generate CA-grade structured Excel
    """
    try:
        session_id = request.data.get('session_id')
        year = int(request.data.get('year'))
        month = int(request.data.get('month'))
        
        # Use shared session utility from gst_auth
        session, error = get_valid_session(session_id)
        if error:
            return Response({"error": error}, status=401)
        
        headers = {
            "Authorization": session.taxpayer_token,
            "x-api-key": settings.SANDBOX_API_KEY,
            "x-api-version": "1.0.0"
        }
        
        # Fetch raw 3B data
        url = f"https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-3b/{year}/{month:02d}"
        status, data = safe_api_call("GET", url, headers=headers)
        
        if status != 200 or not data:
            return Response({"error": "Failed to fetch GSTR-3B data"}, status=400)
        
        gstr3b = data.get("data", {}).get("data", {})
        gstin = gstr3b.get("gstin", session.gstin)
        ret_period = gstr3b.get("ret_period", f"{month:02d}{year}")
        
        # Create workbook
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11, color="1F4E78")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_format = '₹#,##0.00'
        
        month_name = calendar.month_name[month]
        
        # ========== SHEET 1: SUMMARY ==========
        ws = wb.active
        ws.title = "GSTR-3B Summary"
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"GSTR-3B Details - {month_name} {year}"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"GSTIN: {gstin} | Return Period: {ret_period}"
        ws['A2'].font = Font(bold=True, size=11)
        
        row = 4
        
        # ========== TABLE 3.1: OUTWARD SUPPLIES ==========
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 3.1 - OUTWARD SUPPLIES").fill = section_fill
        ws.cell(row, 1).font = section_font
        row += 1
        
        # Headers
        headers_row = ["Particulars", "Taxable Value", "IGST", "CGST", "SGST", "CESS"]
        for col, h in enumerate(headers_row, 1):
            cell = ws.cell(row, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1
        
        sup_details = gstr3b.get("sup_details", {})
        outward_rows = [
            ("3.1(a) Outward taxable supplies", "osup_det"),
            ("3.1(b) Outward taxable supplies (zero rated)", "osup_zero"),
            ("3.1(c) Other outward supplies (Nil rated, exempted)", "osup_nil_exmp"),
            ("3.1(d) Inward supplies liable to reverse charge", "isup_rev"),
            ("3.1(e) Non-GST outward supplies", "osup_nongst"),
        ]
        
        for label, key in outward_rows:
            section = sup_details.get(key, {})
            ws.cell(row, 1, label).border = border
            ws.cell(row, 2, section.get("txval", 0)).number_format = money_format
            ws.cell(row, 2).border = border
            ws.cell(row, 3, section.get("iamt", 0)).number_format = money_format
            ws.cell(row, 3).border = border
            ws.cell(row, 4, section.get("camt", 0)).number_format = money_format
            ws.cell(row, 4).border = border
            ws.cell(row, 5, section.get("samt", 0)).number_format = money_format
            ws.cell(row, 5).border = border
            ws.cell(row, 6, section.get("csamt", 0)).number_format = money_format
            ws.cell(row, 6).border = border
            row += 1
        
        row += 1
        
        # ========== TABLE 4: ITC DETAILS ==========
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 4 - ELIGIBLE ITC").fill = section_fill
        ws.cell(row, 1).font = section_font
        row += 1
        
        # Headers
        for col, h in enumerate(headers_row, 1):
            cell = ws.cell(row, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1
        
        itc_elg = gstr3b.get("itc_elg", {})
        itc_avl = itc_elg.get("itc_avl", [])
        
        itc_type_labels = {
            "IMPG": "4(A)(1) Import of goods",
            "IMPS": "4(A)(2) Import of services",
            "ISRC": "4(A)(3) Inward supplies liable to RCM",
            "ISD": "4(A)(4) Inward supplies from ISD",
            "OTH": "4(A)(5) All other ITC"
        }
        
        for itc_item in itc_avl:
            ty = itc_item.get("ty", "")
            label = itc_type_labels.get(ty, f"4(A) {ty}")
            ws.cell(row, 1, label).border = border
            ws.cell(row, 2, "-").border = border  # No taxable value for ITC
            ws.cell(row, 3, itc_item.get("iamt", 0)).number_format = money_format
            ws.cell(row, 3).border = border
            ws.cell(row, 4, itc_item.get("camt", 0)).number_format = money_format
            ws.cell(row, 4).border = border
            ws.cell(row, 5, itc_item.get("samt", 0)).number_format = money_format
            ws.cell(row, 5).border = border
            ws.cell(row, 6, itc_item.get("csamt", 0)).number_format = money_format
            ws.cell(row, 6).border = border
            row += 1
        
        # ITC Net
        itc_net = itc_elg.get("itc_net", {})
        ws.cell(row, 1, "4(C) Net ITC Available").font = Font(bold=True)
        ws.cell(row, 1).border = border
        ws.cell(row, 2, "-").border = border
        ws.cell(row, 3, itc_net.get("iamt", 0)).number_format = money_format
        ws.cell(row, 3).border = border
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 4, itc_net.get("camt", 0)).number_format = money_format
        ws.cell(row, 4).border = border
        ws.cell(row, 4).font = Font(bold=True)
        ws.cell(row, 5, itc_net.get("samt", 0)).number_format = money_format
        ws.cell(row, 5).border = border
        ws.cell(row, 5).font = Font(bold=True)
        ws.cell(row, 6, itc_net.get("csamt", 0)).number_format = money_format
        ws.cell(row, 6).border = border
        ws.cell(row, 6).font = Font(bold=True)
        row += 2
        
        # ========== TABLE 6: TAX PAYMENT ==========
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 6 - TAX PAYMENT").fill = section_fill
        ws.cell(row, 1).font = section_font
        row += 1
        
        tx_pmt = gstr3b.get("tx_pmt", {})
        net_tax_pay = tx_pmt.get("net_tax_pay", [])
        
        pay_headers = ["Description", "IGST", "CGST", "SGST", "CESS", "Interest"]
        for col, h in enumerate(pay_headers, 1):
            cell = ws.cell(row, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1
        
        for item in net_tax_pay:
            desc = item.get("tran_desc", "")
            ws.cell(row, 1, desc).border = border
            ws.cell(row, 2, item.get("igst", {}).get("tx", 0)).number_format = money_format
            ws.cell(row, 2).border = border
            ws.cell(row, 3, item.get("cgst", {}).get("tx", 0)).number_format = money_format
            ws.cell(row, 3).border = border
            ws.cell(row, 4, item.get("sgst", {}).get("tx", 0)).number_format = money_format
            ws.cell(row, 4).border = border
            ws.cell(row, 5, item.get("cess", {}).get("tx", 0)).number_format = money_format
            ws.cell(row, 5).border = border
            ws.cell(row, 6, item.get("igst", {}).get("intr", 0)).number_format = money_format
            ws.cell(row, 6).border = border
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 45
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
        
        ws.freeze_panes = 'A4'
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="GSTR3B_Details_{gstin}_{month_name}_{year}.xlsx"'
        )
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)