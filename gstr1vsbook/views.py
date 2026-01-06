from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.http import HttpResponse
from django.conf import settings
import pandas as pd
import io

from .serializers import GSTR1ReconciliationRequestSerializer
from .services import GSTR1ReconciliationService
from gst_auth.utils import get_valid_session, safe_api_call
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class GSTR1ReconciliationAPIView(APIView):
    """
    POST: Upload Excel file and get GSTR-1 reconciliation results.
    Uses unified session from gst_auth for authentication.
    """
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request, *args, **kwargs):
        serializer = GSTR1ReconciliationRequestSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        file = data["file"]
        file_bytes = file.read()
        session_id = data.get("session_id")
        
        # Validate session using unified auth from gst_auth
        session, error = get_valid_session(session_id)
        if error:
            return Response({
                "success": False,
                "error": error
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            # Initialize service with credentials from session
            service = GSTR1ReconciliationService(
                api_key=settings.SANDBOX_API_KEY,
                access_token=session.taxpayer_token
            )
            
            results = service.run(
                file_bytes=file_bytes,
                session_id=str(session_id),
                reco_type=data["reco_type"],
                year=data["year"],
                month=data.get("month"),
                quarter=data.get("quarter")
            )
            
            # 1. Fetch Party Name
            party_name = fetch_party_name(session.gstin, session.taxpayer_token) or session.username

            # 2. Convert DataFrames to JSON-serializable dicts
            response_data = {}
            for section, val in results.items():
                if isinstance(val, pd.DataFrame):
                    response_data[section] = val.to_dict(orient="records")
                else:
                    response_data[section] = val # Monthly summary is already a list

            return Response({
                "success": True,
                "data": response_data,
                "session_info": {
                    "party_name": party_name,
                    "gstin": session.gstin,
                    "reco_type": data["reco_type"],
                    "year": data["year"],
                    "month": data.get("month"),
                    "quarter": data.get("quarter")
                }
            }, status=status.HTTP_200_OK)
        
        except ValueError as e:
            return Response({
                "success": False,
                "error": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                "success": False,
                "error": f"Internal error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GSTR1ExcelDownloadAPIView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            results = request.data.get('results', {})
            summary_data = results.get('summary', [])
            if not summary_data:
                return Response({'error': 'No summary data provided'}, status=400)
                
            username = request.data.get('username', '') 
            gstin = request.data.get('gstin', '')
            year = request.data.get('year', '')
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            
            # Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Header Info
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"Username: {username} | GSTIN: {gstin} | FY: {year}"
            ws['A1'].font = Font(bold=True)

            # Labels for rows
            particulars = []
            if summary_data and len(summary_data) > 0 and 'rows' in summary_data[0]:
                particulars = [r['particular'] for r in summary_data[0]['rows']]
            
            if not particulars:
                return Response({'error': 'Summary data format invalid or empty rows'}, status=400)
            
            # Start writing headers
            ws.cell(row=3, column=1, value="Particular").font = Font(bold=True)
            ws.cell(row=3, column=1).fill = header_fill
            ws.cell(row=3, column=1).font = header_font
            ws.cell(row=3, column=1).border = border
            
            col_idx = 2
            for m_block in summary_data:
                month_name = m_block['month']
                ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+2)
                cell = ws.cell(row=3, column=col_idx, value=month_name)
                cell.font = header_font
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = center_align
                cell.border = border
                
                # Sub-headers
                h1 = ws.cell(row=4, column=col_idx, value="Books")
                h2 = ws.cell(row=4, column=col_idx+1, value="GSTR-1")
                h3 = ws.cell(row=4, column=col_idx+2, value="Diff")
                for h in [h1, h2, h3]:
                    h.font = Font(bold=True, size=8); h.border = border
                    h.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                col_idx += 4
                
            # Write Particulars
            for i, part in enumerate(particulars, 5):
                cell = ws.cell(row=i, column=1, value=part)
                cell.font = Font(bold=True, size=9); cell.border = border

            # Write Data
            col_idx = 2
            for m_block in summary_data:
                for i, row in enumerate(m_block['rows'], 5):
                    c1 = ws.cell(row=i, column=col_idx, value=row['v1'])
                    c2 = ws.cell(row=i, column=col_idx+1, value=row['v2'])
                    c3 = ws.cell(row=i, column=col_idx+2, value=row['diff'])
                    for c in [c1, c2, c3]:
                        c.border = border; c.number_format = '#,##0.00'; c.font = Font(size=9)
                    if abs(row['diff']) > 1.0:
                        c3.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                    else:
                        c3.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                col_idx += 4
                
            # Adjust widths
            ws.column_dimensions['A'].width = 25
            for i in range(2, col_idx):
                ws.column_dimensions[get_column_letter(i)].width = 12

            # Add detail sheets
            sections = ["B2B", "B2CL", "B2CS", "EXP", "CDNR"]
            header_map = {
                "Taxable_BOOKS": "Books Taxable", "IGST_BOOKS": "Books IGST", "CGST_BOOKS": "Books CGST", "SGST_BOOKS": "Books SGST",
                "Taxable_PORTAL": "Portal Taxable", "IGST_PORTAL": "Portal IGST", "CGST_PORTAL": "Portal CGST", "SGST_PORTAL": "Portal SGST",
                "Taxable_DIFF": "Diff Taxable", "IGST_DIFF": "Diff IGST", "CGST_DIFF": "Diff CGST", "SGST_DIFF": "Diff SGST"
            }

            for section in sections:
                records = results.get(section, [])
                if records:
                    detail_ws = wb.create_sheet(title=section)
                    df = pd.DataFrame(records)
                    
                    # Rename columns for display
                    display_cols = [header_map.get(c, c) for c in df.columns]
                    
                    # Header Style
                    for c_idx, col_name in enumerate(display_cols, 1):
                        cell = detail_ws.cell(row=1, column=c_idx, value=col_name)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = center_align
                        
                    # Data and Formatting
                    for r_idx, row_values in enumerate(df.values, 2):
                        for c_idx, value in enumerate(row_values, 1):
                            cell = detail_ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.border = border
                            
                            col_name = df.columns[c_idx-1]
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
                                
                            # Highlight differences
                            if "_DIFF" in col_name and isinstance(value, (int, float)):
                                if abs(value) > 1.0:
                                    cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                                else:
                                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                    # Auto-adjust column widths
                    for i, col in enumerate(display_cols, 1):
                        max_length = len(str(col)) + 4
                        detail_ws.column_dimensions[get_column_letter(i)].width = max_length

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="GSTR1_Reconciliation_{gstin}_{year}.xlsx"'
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"success": False, "error": f"Download failed: {str(e)}"}, status=500)


def fetch_party_name(gstin, token):
    """Fetch Legal/Trade Name from Sandbox"""
    try:
        from django.conf import settings
        status, data = safe_api_call(
            "GET",
            f"https://api.sandbox.co.in/gst/compliance/tax-payer/details?gstin={gstin}",
            headers={
                "x-api-version": "1.0.0",
                "Authorization": token,
                "x-api-key": settings.SANDBOX_API_KEY
            }
        )
        if status == 200:
            return data.get("data", {}).get("tradeNam") or data.get("data", {}).get("lgnm")
    except:
        pass
    return None