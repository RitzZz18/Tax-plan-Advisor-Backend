import requests
import pandas as pd
from datetime import datetime
import time
import uuid
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_BASE = "https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-1"

UNWANTED_PREFIXES = (
    "status",
    "error",
    "timestamp",
    "transaction",
    "data.b2b.inv.itms.num",
    "data.b2b.inv.flag",
    "data.b2b.inv.updby",
    "data.b2b.inv.cflag","data.cdnr.nt.cflag",
    "data.b2b.inv.chksum","data.b2ba.inv.chksum","data.b2cs.chksum","data.b2csa.chksum","data.b2csa.chksum","data.cdnr.chksum","data.cdnra.nt.chksum","data.hsn.chksum","data.cdnr.nt.chksum","data.exp.inv.chksum","data.doc_issue.chksum",
    "data.b2b.inv.itms.num","data.b2ba.inv.itms.num","data.cdnra.nt.itms.num",
    "data.b2b.cfs","data.b2ba.cfs","data.cdnra.cfs","data.cdnr.cfs",
    "data.b2ba.inv.itms.num","data.b2ba.inv.cflag","data.cdnr.nt.itms.num",
    "data.b2ba.inv.flag","data.b2cs.flag","data.b2csa.flag","data.cdnr.nt.flag","data.cdnra.nt.flag","data.cdnra.nt.d_flag","data.cdnr.nt.d_flag","data.cdnra.nt.cflag","data.exp.inv.flag",
    "data.b2ba.inv.updby","data.cdnr.inv.updby","data.cdnra.nt.updby","data.cdnr.nt.updby",
    "data.cdnr.nt.ntty","data.cdnra.nt.ntty",
    "data.hsn.data.uqc",
    "data.doc_issue.flag",
    "data.hsn.flag",
    "data.doc_issue.doc_det.docs.num","data.doc_issue.doc_det.doc_num",
)

VALUE_MAPPING = {
    "Invoice Type": {"R": "Regular"},
    "Reverse Charge": {"Y": "Yes", "N": "No"},
}

ENDPOINTS = {
    "summary": "",  # Base endpoint for summary
    "b2b": "b2b",
    "b2ba": "b2ba",
    "b2cl": "b2cl",
    "b2cla": "b2cla",
    "b2cs": "b2cs",
    "b2csa": "b2csa",
    "cdnr": "cdnr",
    "cdnra": "cdnra",
    "cdnur": "cdnur",
    "cdnura": "cdnura",
    "exp": "exp",
    "nil": "nil",
    "hsn": "hsn",
    "docs": "doc-issue",
    "at": "at",
    "ata": "ata"
}

OUTPUT_COLUMNS = [
    "Return Period", "Filing Status", 
    "Original Invoice Number", "Original Invoice Date", # Amendments only
    "Invoice Number", "Invoice Date", "Invoice Value",
    "Place of Supply", "Reverse Charge", "E-Commerce GSTIN", "Invoice Type",
    "Applicable % of Tax Rate", "GSTIN/UIN of Recipient", "Receiver Name", "Rate",
    "Taxable Value", "Tax Amount", "IGST Amount", "CGST Amount", "SGST Amount",
    "CESS Amount", "IBN", "Generation Date", "Source Type",
    # Docs Sheet Columns
    "Nature of Document", "Sr. No. From", "Sr. No. To", "Total Number", "Cancelled", "Net issued",
    # HSN Sheet Columns
    "HSN", "Description", "UQC", "Total Quantity", "Total Value",
    # Summary Sheet Columns
    "Section name", "Number of documents", "Total Amount"
]

COLUMN_MAPPING = {
    "ctin": "GSTIN/UIN of Recipient",
    "cname": "Receiver Name",
    "oinum": "Original Invoice Number",
    "oidt": "Original Invoice Date",
    "inum": "Invoice Number",
    "idt": "Invoice Date",
    "val": "Invoice Value",
    "pos": "Place of Supply",
    "rchrg": "Reverse Charge",
    "etin": "E-Commerce GSTIN",
    "inv_typ": "Invoice Type",
    "rt": "Rate",
    "txval": "Taxable Value",
    "iamt": "IGST Amount",
    "camt": "CGST Amount",
    "samt": "SGST Amount",
    "csamt": "CESS Amount",
    "nt_num": "Invoice Number",
    "nt_dt": "Invoice Date",
    "ont_num": "Original Invoice Number",
    "ont_dt": "Original Invoice Date",
    "omon": "Original Month",
    "typ": "Original Month",
    "sply_ty": "Supply Type",
    "exp_typ": "Export Type",
    "doc_desc": "Nature of Document", 
    "from": "Sr. No. From",
    "to": "Sr. No. To",
    "totnum": "Total Number",
    "cancel": "Cancelled",
    "net_issue": "Net issued",
    
    # HSN Mappings
    "hsn_sc": "HSN",
    "desc": "Description",
    "uqc": "UQC",
    "qty": "Total Quantity",
 
    # Summary Mappings
    "sec_nm": "Section name",
    "ttl_doc": "Number of documents",
    "ttl_tax": "Taxable Value",
    "ttl_igst": "IGST Amount",
    "ttl_cgst": "CGST Amount",
    "ttl_sgst": "SGST Amount",
    "ttl_cess": "CESS Amount",
    "ttl_val": "Total Amount"
}

def get_fy_months(fy):
    y = int(fy.split("-")[0])
    now = datetime.now()
    months = []
    for m in range(4, 13):
        if y < now.year or (y == now.year and m <= now.month):
            months.append((y, f"{m:02d}"))
    for m in range(1, 4):
        if y + 1 < now.year or (y + 1 == now.year and m <= now.month):
            months.append((y + 1, f"{m:02d}"))
    return months

def get_quarterly_months(fy, quarter):
    y = int(fy.split("-")[0])
    q = int(quarter)
    if q == 1: return [(y, "04"), (y, "05"), (y, "06")]
    elif q == 2: return [(y, "07"), (y, "08"), (y, "09")]
    elif q == 3: return [(y, "10"), (y, "11"), (y, "12")]
    else: return [(y+1, "01"), (y+1, "02"), (y+1, "03")]

def get_monthly_period(year, month):
    return [(int(year), month)]

def fetch_data(api_key, access_token, endpoint, year, month, retries=5):
    url = f"{API_BASE}/{endpoint}/{year}/{month}" if endpoint else f"{API_BASE}/{year}/{month}"
    headers = {"x-api-key": api_key, "Authorization": access_token, "accept": "application/json"}
    
    last_error = "Unknown error"
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code == 200:
                return r.json().get("data", {})
            elif r.status_code == 503 or r.status_code == 429: # Service busy or Rate limit
                last_error = f"Server busy (Status {r.status_code})"
                if attempt < retries - 1:
                    time.sleep(2 * (attempt + 1))
                    continue
            else:
                last_error = f"API returned status {r.status_code}"
        except requests.exceptions.ReadTimeout:
            last_error = "Request timed out"
            if attempt < retries - 1:
                time.sleep(2)
                continue
        except Exception as e:
            last_error = str(e)
            
    # If we reached here, all retries failed
    raise Exception(f"Failed to fetch {endpoint or 'summary'} for {month}/{year}: {last_error}")

def flatten_json(data, parent="", rows=None):
    if rows is None:
        rows = [{}]
    if isinstance(data, dict):
        for k, v in data.items():
            rows = flatten_json(v, f"{parent}.{k}" if parent else k, rows)
    elif isinstance(data, list):
        new_rows = []
        for item in data:
            for r in rows:
                base = r.copy()
                new_rows.extend(flatten_json(item, parent, [base]))
        rows = new_rows
    else:
        for r in rows:
            r[parent] = data
    return rows

def clean_dataframe(df, sheet_name=""):
    if df.empty:
        return df
    df = df[[c for c in df.columns if not c.startswith(UNWANTED_PREFIXES)]]
    renamed_cols = {}
    for col in df.columns:
        for key, value in COLUMN_MAPPING.items():
            if key in col.split('.'):
                renamed_cols[col] = value
                break
    df = df.rename(columns=renamed_cols)
    if sheet_name == "hsn" and "Invoice Value" in df.columns:
        df = df.rename(columns={"Invoice Value": "Total Value"})
    df = df.loc[:, ~df.columns.duplicated()]
    for col, mapping in VALUE_MAPPING.items():
        if col in df.columns:
            df[col] = df[col].map(mapping).fillna(df[col])
    tax_cols = ["IGST Amount", "CGST Amount", "SGST Amount", "CESS Amount"]
    existing_tax_cols = [c for c in tax_cols if c in df.columns]
    if existing_tax_cols:
        for c in existing_tax_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        df["Tax Amount"] = df[existing_tax_cols].sum(axis=1)
    present_cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in present_cols]
    df = df[present_cols + remaining_cols]
    return df.reset_index(drop=True)

def generate_excel(gstin, api_key, access_token, download_type, fy, quarter, year, month):
    if download_type == "fy":
        months_list = get_fy_months(fy)
        period_label = fy
    elif download_type == "quarterly":
        months_list = get_quarterly_months(fy, quarter)
        period_label = f"{fy}_Q{quarter}"
    else:
        months_list = get_monthly_period(year, month)
        period_label = f"{month}{year}"
    
    sheets = {k: [] for k in ENDPOINTS}
    
    # Prepare all tasks for parallel execution
    tasks = []
    for yr, mn in months_list:
        for sheet, endpoint in ENDPOINTS.items():
            tasks.append((sheet, endpoint, yr, mn))
    
    errors = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_task = {
            executor.submit(fetch_data, api_key, access_token, t[1], t[2], t[3]): t 
            for t in tasks
        }
        
        for future in as_completed(future_to_task):
            sheet, endpoint, yr, mn = future_to_task[future]
            try:
                data = future.result()
                if data:
                    rows = flatten_json(data)
                    for r in rows:
                        r["Month"] = mn
                        r["Return Period"] = f"{mn}{yr}"
                        r["Filing Status"] = "FILED"
                        r["Source Type"] = "Manual"
                    sheets[sheet].extend(rows)
            except Exception as e:
                errors.append(str(e))
    
    # If any error occurred, we stop and notify the user
    if errors:
        # Join errors and show the first unique ones to keep it readable
        unique_errors = list(set(errors))[:3]
        raise Exception("Download aborted to prevent incomplete data: " + " | ".join(unique_errors))
    
    # Generate Excel in memory
    output = io.BytesIO()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = str(uuid.uuid4())[:8]
    filename = f"GSTR1_{gstin}_{period_label}_{timestamp}_{unique_id}.xlsx"
    
    header_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    blue_diff_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True, name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_count = 0
        for sheet, rows in sheets.items():
            if not rows: continue
            df = pd.DataFrame(rows)
            df = clean_dataframe(df, sheet_name=sheet)
            df.to_excel(writer, sheet_name=sheet, index=False)
            sheet_count += 1
            
            if sheet in writer.sheets:
                worksheet = writer.sheets[sheet]
                for col_num, value in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = blue_diff_fill if "Original" in str(value) else header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    worksheet.column_dimensions[get_column_letter(col_num)].width = 20
        
        if sheet_count == 0:
            pd.DataFrame(columns=OUTPUT_COLUMNS).to_excel(writer, sheet_name="No Data", index=False)
    
    output.seek(0)
    return output, filename
