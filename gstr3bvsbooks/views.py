from django.http import HttpResponse
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

from gst_auth.utils import get_valid_session, safe_api_call


@api_view(['POST'])
@permission_classes([AllowAny])
def reconciliation(request):
    """
    GSTR-3B vs Books Reconciliation.
    Uses unified session from gst_auth for authentication.
    """
    session_id = request.data.get('session_id')
    reco_type = request.data.get('reco_type')
    year = request.data.get('year')
    month = request.data.get('month')
    quarter = request.data.get('quarter')
    
    if not session_id:
        return Response({'error': 'Session ID required'}, status=400)
        
    if not all([reco_type, year]):
        return Response({'error': 'Missing required fields'}, status=400)
    
    # Validate session using unified auth
    session, error = get_valid_session(session_id)
    if error:
        return Response({'error': error}, status=401)
    
    # Handle file upload
    if 'file' not in request.FILES:
        return Response({'error': 'No file uploaded'}, status=400)
    
    try:
        file = request.FILES['file']
        df = pd.read_excel(file)
        
        # Get months to process
        months_list = get_months(reco_type, int(year), int(month) if month else None, quarter)
        
        # 1. Fetch Party Name
        party_name = fetch_party_name(session.gstin, session.taxpayer_token) or session.username

        # 2. Process Books Data
        norm_df = normalize_helper_data(df, months_list)
        books_monthly = calculate_books_monthly(norm_df, months_list) # Returns { "YYYY-MM": { "3.1(a)": ... } }

        # 3. Process Portal Data
        portal_monthly = fetch_portal_monthly(months_list, session.taxpayer_token)

        # 4. Process Difference Data
        diff_monthly, status_monthly = calculate_diff_monthly(books_monthly, portal_monthly)

        # 5. Format for Frontend
        final_report = []
        sorted_months = sorted(books_monthly.keys())
        
        # Rows to include (skipping 3.1(d))
        particular_mapping = [
            ("3.1.a Taxable Value", "3.1(a)", "taxable"),
            ("3.1.a IGST", "3.1(a)", "igst"),
            ("3.1.a CGST", "3.1(a)", "cgst"),
            ("3.1.a SGST", "3.1(a)", "sgst"),
            ("3.1.b Exports Taxable", "3.1(b)", "taxable"),
            ("3.1.b Exports IGST", "3.1(b)", "igst"),
            ("3.1.c Nil/Exempt", "3.1(c)", "taxable"),
            ("3.1.e Non-GST", "3.1(e)", "taxable"),
        ]

        # Helper for month name
        def get_month_display(m_str):
            # m_str is "YYYY-MM"
            import datetime
            dt = datetime.datetime.strptime(m_str, "%Y-%m")
            return dt.strftime("%b %Y")

        for m in sorted_months:
            month_rows = []
            m_status = "MATCHED"
            
            for part_label, sec, field in particular_mapping:
                v1 = books_monthly[m].get(sec, {}).get(field, 0)
                v2 = portal_monthly.get(m, {}).get(sec, {}).get(field, 0)
                diff = v1 - v2
                
                if abs(diff) > 1.0:
                    m_status = "MISMATCHED"
                
                month_rows.append({
                    "particular": part_label,
                    "v1": v1,
                    "v2": v2,
                    "diff": diff
                })
            
            final_report.append({
                "month": get_month_display(m),
                "month_key": m,
                "status": m_status,
                "rows": month_rows
            })
            
        return Response({
            'status': 'success',
            'message': 'Reconciliation completed',
            'data': final_report,
            'session_info': {
                'party_name': party_name,
                'gstin': session.gstin,
                'reco_type': reco_type,
                'year': year,
                'month': month,
                'quarter': quarter
            }
        })
                
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'Processing error: {str(e)}'}, status=500)


def fetch_party_name(gstin, token):
    """Fetch Legal/Trade Name from Sandbox"""
    try:
        status, data = safe_api_call(
            "GET",
            f"https://api.sandbox.co.in/gst/compliance/tax-payer/details?gstin={gstin}",
            headers={
                "x-api-version": "1.0.0",
                "Authorization": token,
                "x-api-key": settings.SANDBOX_API_KEY
            }
        )
        if status == 200:
            return data.get("data", {}).get("tradeNam") or data.get("data", {}).get("lgnm")
    except:
        pass
    return None


def get_months(reco_type, year, month=None, quarter=None):
    if reco_type == "MONTHLY":
        return [(year, month)]
    if reco_type == "QUARTERLY":
        q_map = {"Q1": [4,5,6], "Q2": [7,8,9], "Q3": [10,11,12], "Q4": [1,2,3]}
        return [(year if m >= 4 else year + 1, m) for m in q_map[quarter]]
    if reco_type == "FY":
        return [(year, m) for m in range(4,13)] + [(year+1, m) for m in range(1,4)]


def normalize_helper_data(df, valid_months):
    """
    Normalize the input dataframe (Books) to a standard format.
    Handles columns from the "New Template" (same as GSTR1).
    Filters data by the valid_months list [(year, month), ...].
    """
    # 1. Clean Column Names
    df.columns = df.columns.astype(str).str.strip()
    
    # 2. Date Filtering
    if "Date" in df.columns and valid_months:
        try:
            # Convert to datetime, handling errors gracefully
            df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors='coerce')
            
            # Filter rows where (year, month) matches valid_months
            mask = df["Date"].apply(lambda x: (x.year, x.month) in valid_months if pd.notnull(x) else False)
            df = df[mask].copy()
        except Exception as e:
            # Fallback or log if needed, though 'coerce' handles most
            pass

    if df.empty:
        return pd.DataFrame()
    
    # 3. Ensure Numeric Columns exist and are float
    numeric_cols = ["Taxable", "Export_Taxable", "SEZ_Taxable", "Nil_Rated", 
                    "Exempt", "Non_GST", "IGST", "CGST", "SGST", "Cess"]
    
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            
    # Handle RCM column
    if "Is_RCM" not in df.columns:
        df["Is_RCM"] = "N"
    df["Is_RCM"] = df["Is_RCM"].fillna("N").astype(str).str.upper().str.strip()

    normalized_rows = []
    
    for _, r in df.iterrows():
        # Logic adapted from gstr1vsbook services
        
        # Calculate derived values
        igst, cgst, sgst = r["IGST"], r["CGST"], r["SGST"]
        tax = igst + cgst + sgst
        
        # Determine Category and specific Taxable Value
        # Priority: RCM > Export > SEZ > Nil/Exempt > Non-GST > Standard (B2B/B2C)
        
        sup_cat = "DOM"
        taxable_val = 0.0
        
        if r["Is_RCM"] == "Y":
            sup_cat = "RCM"
            # In RCM, the 'Taxable' column usually holds the value, or user might put it in others.
            # We'll take the sum of all taxable components to be safe, or just 'Taxable'.
            # gstr1vsbook sums everything:
            taxable_val = r["Taxable"] + r["Export_Taxable"] + r["SEZ_Taxable"] + \
                          r["Nil_Rated"] + r["Exempt"] + r["Non_GST"]
                          
        elif r["Export_Taxable"] > 0:
            sup_cat = "EXPWP" if tax > 0 else "EXPWOP"
            taxable_val = r["Export_Taxable"]
            
        elif r["SEZ_Taxable"] > 0:
            sup_cat = "SEZWP" if tax > 0 else "SEZWOP"
            taxable_val = r["SEZ_Taxable"]
            
        elif r["Nil_Rated"] > 0 or r["Exempt"] > 0:
            sup_cat = "NIL"
            taxable_val = r["Nil_Rated"] + r["Exempt"]
            
        elif r["Non_GST"] > 0:
            sup_cat = "NON_GST"
            taxable_val = r["Non_GST"]
            
        else:
            # Domestic (B2B, B2C, etc.) -> Mapped to 3.1(a)
            sup_cat = "DOM"
            taxable_val = r["Taxable"]

        normalized_rows.append({
            "SUP_CAT": sup_cat,
            "Taxable": taxable_val,
            "IGST": igst,
            "CGST": cgst,
            "SGST": sgst,
            "Is_RCM": r["Is_RCM"],
            "Year": r["Date"].year if "Date" in df.columns and pd.notnull(r["Date"]) else 0,
            "Month": r["Date"].month if "Date" in df.columns and pd.notnull(r["Date"]) else 0
        })

    return pd.DataFrame(normalized_rows)


def calculate_books_monthly(norm_df, months_list):
    """
    Returns { "YYYY-MM": { "3.1(a)": {metrics}, ... } }
    """
    sections = ["3.1(a)", "3.1(b)", "3.1(c)", "3.1(d)", "3.1(e)"]
    def init_metrics():
        return {"taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "tax": 0.0}

    # Initialize all requested months with 0
    monthly_data = {}
    for y, m in months_list:
        m_key = f"{y}-{m:02d}"
        monthly_data[m_key] = {k: init_metrics() for k in sections}

    if norm_df.empty:
        return monthly_data

    for _, r in norm_df.iterrows():
        # Skip if no date
        if r["Year"] == 0 or r["Month"] == 0:
            continue
            
        m_key = f"{int(r['Year'])}-{int(r['Month']):02d}"
        if m_key not in monthly_data:
            # Maybe data outside range, ignore or add? 
            # Given we filtered norm_df by months_list, this updates only valid months
            continue

        tax = r["IGST"] + r["CGST"] + r["SGST"]
        
        # Mapping Logic
        key = None
        if r["SUP_CAT"] == "RCM": key = "3.1(d)"
        elif r["SUP_CAT"] in ("EXPWP", "EXPWOP", "SEZWP", "SEZWOP"): key = "3.1(b)"
        elif r["SUP_CAT"] == "NIL": key = "3.1(c)"
        elif r["SUP_CAT"] == "NON_GST": key = "3.1(e)"
        elif r["SUP_CAT"] == "DOM": key = "3.1(a)"
            
        if key:
            monthly_data[m_key][key]["taxable"] += r["Taxable"]
            monthly_data[m_key][key]["igst"] += r["IGST"]
            monthly_data[m_key][key]["cgst"] += r["CGST"]
            monthly_data[m_key][key]["sgst"] += r["SGST"]
            monthly_data[m_key][key]["tax"] += tax
    
    return monthly_data


def fetch_portal_monthly(months_list, taxpayer_access_token):
    sections = ["3.1(a)", "3.1(b)", "3.1(c)", "3.1(d)", "3.1(e)"]
    def init_metrics():
        return {"taxable":0,"igst":0,"cgst":0,"sgst":0,"tax":0}

    monthly_data = {}
    
    for y, m in months_list:
        m_key = f"{y}-{m:02d}"
        monthly_data[m_key] = {k: init_metrics() for k in sections}
        
        status_code, response_data = safe_api_call(
            "GET",
            f"https://api.sandbox.co.in/gst/compliance/tax-payer/gstrs/gstr-3b/{y}/{str(m).zfill(2)}",
            headers={
                "x-api-version": "1.0.0",
                "Authorization": taxpayer_access_token,
                "x-api-key": settings.SANDBOX_API_KEY
            }
        )
        
        if status_code != 200:
            continue
            
        sup = response_data.get("data", {}).get("data", {}).get("sup_details", {})
        
        # Helper to process a section
        def process_sec(sec_key, source_dict):
            if not source_dict: return
            txval = source_dict.get("txval", 0)
            iamt = source_dict.get("iamt", 0)
            camt = source_dict.get("camt", 0)
            samt = source_dict.get("samt", 0)
            tax = iamt + camt + samt
            
            monthly_data[m_key][sec_key]["taxable"] += txval
            monthly_data[m_key][sec_key]["igst"] += iamt
            monthly_data[m_key][sec_key]["cgst"] += camt
            monthly_data[m_key][sec_key]["sgst"] += samt
            monthly_data[m_key][sec_key]["tax"] += tax

        process_sec("3.1(a)", sup.get("osup_det"))
        process_sec("3.1(b)", sup.get("osup_zero"))
        process_sec("3.1(c)", sup.get("osup_nil_exmp")) 
        process_sec("3.1(d)", sup.get("isup_rev"))
        process_sec("3.1(e)", sup.get("osup_nongst"))

    return monthly_data


def calculate_diff_monthly(books_monthly, portal_monthly):
    sections = ["3.1(a)", "3.1(b)", "3.1(c)", "3.1(d)", "3.1(e)"]
    def init_metrics(): return {"taxable":0,"igst":0,"cgst":0,"sgst":0,"tax":0}
    
    diff_monthly = {}
    status_monthly = {}
    
    all_months = sorted(list(set(list(books_monthly.keys()) + list(portal_monthly.keys()))))
    
    for m in all_months:
        diff_monthly[m] = {}
        status_monthly[m] = {}
        
        for sec in sections:
            b = books_monthly.get(m, {}).get(sec, init_metrics())
            p = portal_monthly.get(m, {}).get(sec, init_metrics())
            
            d_taxable = b['taxable'] - p['taxable']
            d_igst = b['igst'] - p['igst']
            d_cgst = b['cgst'] - p['cgst']
            d_sgst = b['sgst'] - p['sgst']
            d_tax = b['tax'] - p['tax']
            
            diff_monthly[m][sec] = {
                'taxable': d_taxable, 'igst': d_igst, 'cgst': d_cgst, 'sgst': d_sgst, 'tax': d_tax
            }
            
            # Status
            status = "Matched"
            if (abs(d_taxable) + abs(d_igst) + abs(d_cgst) + abs(d_sgst)) > 1.0:
                 if sec == "3.1(d)" and b['taxable'] == 0 and p['taxable'] > 0:
                     status = "RCM - Purchase Side"
                 else:
                     status = "Mismatch"
            status_monthly[m][sec] = status
            
    return diff_monthly, status_monthly


@api_view(['POST'])
@permission_classes([AllowAny])
def download_excel(request):
    try:
        report_data = request.data.get('results', []) # Now a list of monthly blocks
        if not report_data:
            return Response({'error': 'No results data provided'}, status=400)
            
        username = request.data.get('username', '') 
        gstin = request.data.get('gstin', '')
        reco_type = request.data.get('reco_type', '')
        year = request.data.get('year', '')
        month = request.data.get('month', '')
        quarter = request.data.get('quarter', '')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-3B Reconciliation"
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Header Info
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"Username: {username} | GSTIN: {gstin} | FY: {year}"
        ws['A1'].font = Font(bold=True)

        # Labels for rows
        particulars = [r['particular'] for r in report_data[0]['rows']] if report_data else []
        
        # Start writing headers
        ws.cell(row=3, column=1, value="Particular").font = Font(bold=True)
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=1).border = border
        
        col_idx = 2
        for m_block in report_data:
            month_name = m_block['month']
            # Merge 3 cells for month title
            ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+2)
            cell = ws.cell(row=3, column=col_idx, value=month_name)
            cell.font = header_font
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
            
            # Sub-headers
            h1 = ws.cell(row=4, column=col_idx, value="Books")
            h2 = ws.cell(row=4, column=col_idx+1, value="GSTR-3B")
            h3 = ws.cell(row=4, column=col_idx+2, value="Diff")
            for h in [h1, h2, h3]:
                h.font = Font(bold=True, size=8)
                h.border = border
                h.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            col_idx += 4 # 3 columns data + 1 gap
            
        # Write Particulars
        for i, part in enumerate(particulars, 5):
            cell = ws.cell(row=i, column=1, value=part)
            cell.font = Font(bold=True, size=9)
            cell.border = border

        # Write Data
        col_idx = 2
        for m_block in report_data:
            for i, row in enumerate(m_block['rows'], 5):
                c1 = ws.cell(row=i, column=col_idx, value=row['v1'])
                c2 = ws.cell(row=i, column=col_idx+1, value=row['v2'])
                c3 = ws.cell(row=i, column=col_idx+2, value=row['diff'])
                
                for c in [c1, c2, c3]:
                    c.border = border
                    c.number_format = '#,##0.00'
                    c.font = Font(size=9)
                
                # Highlight diff if mismatch
                if abs(row['diff']) > 1.0:
                    c3.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                else:
                    c3.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    
            col_idx += 4
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for i in range(2, col_idx):
            ws.column_dimensions[get_column_letter(i)].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="GSTR3B_Reconciliation_{gstin}_{year}.xlsx"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'Export failed: {str(e)}'}, status=500)
